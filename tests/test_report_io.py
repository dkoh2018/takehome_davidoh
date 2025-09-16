from unittest.mock import patch, mock_open
from src.report_io import read_movie_ids, read_movie_ids_with_skips, write_excel, write_failed_ids
import os
import csv


def test_read_movie_ids_happy_path():
    csv_content = "ID\n123\n456\n789"
    mock_file = mock_open(read_data=csv_content)

    with patch("builtins.open", mock_file), \
         patch("os.path.exists", return_value=True):
        result = read_movie_ids("movies.csv")
        
    assert result == [123, 456, 789]


def test_read_movie_ids_handles_duplicates_and_blanks():
    csv_content = "ID\n123\n\n456\n123\ninvalid\n789"
    mock_file = mock_open(read_data=csv_content)

    with patch("builtins.open", mock_file), \
         patch("os.path.exists", return_value=True):
        result = read_movie_ids("movies.csv")
        
    assert result == [123, 456, 789]


def test_read_movie_ids_file_not_found():
    with patch("os.path.exists", return_value=False):
        result = read_movie_ids("movies.csv")
        
    assert result == []


def test_read_movie_ids_with_skips_collects_reasons_and_bom_header():
    # Includes BOM in header, non-numeric, blank, duplicate, non-positive
    bom = "\ufeff"
    csv_content = f"{bom}ID\n123\n \n456\n123\n-1\ninvalid\n789\n"
    mock_file = mock_open(read_data=csv_content)

    with patch("builtins.open", mock_file), \
         patch("os.path.exists", return_value=True):
        ids, skipped = read_movie_ids_with_skips("movies.csv")

    assert ids == [123, 456, 789]
    reasons = {(raw, reason) for raw, _, reason in skipped}
    assert ("", "blank") in reasons
    assert ("123", "duplicate") in reasons
    assert ("-1", "non-positive") in reasons
    assert ("invalid", "non-numeric") in reasons


def test_write_failed_ids_writes_expected_rows(tmp_path):
    output = os.path.join(tmp_path, "dead_letter_queue.csv")
    failed_ids = [804143, 9999999]
    skipped_rows = [("399S79", 117, "non-numeric")]

    write_failed_ids(failed_ids, skipped_rows, output)

    with open(output, newline="") as f:
        rows = list(csv.reader(f))

    assert rows[0] == ["ID", "Reason", "RowNumber"]
    assert ["804143", "fetch-failed", ""] in rows
    assert ["9999999", "fetch-failed", ""] in rows
    assert ["399S79", "non-numeric", "117"] in rows


def test_write_excel_formats_action_rows(tmp_path):
    # Minimal MovieRow objects
    from src.schemas import MovieRow
    rows = [
        MovieRow(id=1, title="A", vote_average=7.5, genres=["Action"], is_action=True),
        MovieRow(id=2, title="B", vote_average=6.0, genres=["Drama"], is_action=False),
    ]
    output = os.path.join(tmp_path, "movie_data.xlsx")

    write_excel(rows, output)

    # Verify file exists and basic structure via openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb.active
    assert ws.title == "Movies"
    assert [cell.value for cell in ws[1]] == ["ID", "Title", "Vote Average", "Genres"]
    # Row 2 (Action) should be bold + red fill
    row2 = ws[2]
    assert all(cell.font.bold for cell in row2)
    assert all(cell.fill.start_color.rgb in ("00FFC7CE", "FFC7CE") for cell in row2)
    # Row 3 (non-Action) should be normal
    row3 = ws[3]
    assert not any(cell.font.bold for cell in row3)