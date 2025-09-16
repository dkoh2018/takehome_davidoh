import csv
import os
import logging
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError

from src.schemas import MovieRow


def read_movie_ids(csv_path: str) -> List[int]:
    movie_ids = []
    seen_ids = set()

    if not os.path.exists(csv_path):
        logging.error(f"Input file not found at: {csv_path}")
        return []

    with open(csv_path, mode="r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)

        # Normalize potential BOM in header (helps with mocks that bypass encoding handling)
        if reader.fieldnames:
            reader.fieldnames = [
                (name.lstrip("\ufeff") if isinstance(name, str) else name)
                for name in reader.fieldnames
            ]

        if "ID" not in reader.fieldnames:
            logging.error(f"Input file {csv_path} is missing the required 'ID' header.")
            return []

        for row_num, row in enumerate(reader, start=2):
            raw_value = row.get("ID")
            id_value = (raw_value or "").strip()

            if not id_value:
                # Skip blank rows quietly in this variant
                continue

            try:
                movie_id = int(id_value)
            except ValueError:
                logging.warning(
                    f"Skipping non-numeric ID '{id_value}' on row {row_num}."
                )
                continue

            if movie_id <= 0:
                logging.warning(
                    f"Skipping non-positive ID '{id_value}' on row {row_num}."
                )
                continue

            if movie_id in seen_ids:
                # Silently drop duplicates in this variant
                continue

            seen_ids.add(movie_id)
            movie_ids.append(movie_id)

    return movie_ids


def read_movie_ids_with_skips(
    csv_path: str,
) -> Tuple[List[int], List[Tuple[str, int, str]]]:
    movie_ids: List[int] = []
    seen_ids = set()
    skipped: List[Tuple[str, int, str]] = []  # (raw_value, row_number, reason)

    if not os.path.exists(csv_path):
        logging.error(f"Input file not found at: {csv_path}")
        return [], skipped

    with open(csv_path, mode="r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)

        # Normalize potential BOM in header (helps with mocks that bypass encoding handling)
        if reader.fieldnames:
            reader.fieldnames = [
                (name.lstrip("\ufeff") if isinstance(name, str) else name)
                for name in reader.fieldnames
            ]

        if "ID" not in reader.fieldnames:
            logging.error(f"Input file {csv_path} is missing the required 'ID' header.")
            return [], skipped

        for row_num, row in enumerate(reader, start=2):
            raw_value = row.get("ID")
            id_value = (raw_value or "").strip()

            if not id_value:
                skipped.append((id_value, row_num, "blank"))
                continue

            try:
                movie_id = int(id_value)
            except ValueError:
                logging.warning(
                    f"Skipping non-numeric ID '{id_value}' on row {row_num}."
                )
                skipped.append((id_value, row_num, "non-numeric"))
                continue

            if movie_id <= 0:
                logging.warning(
                    f"Skipping non-positive ID '{id_value}' on row {row_num}."
                )
                skipped.append((id_value, row_num, "non-positive"))
                continue

            if movie_id in seen_ids:
                skipped.append((id_value, row_num, "duplicate"))
                continue

            seen_ids.add(movie_id)
            movie_ids.append(movie_id)

    return movie_ids, skipped


def write_excel(rows: List[MovieRow], output_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movies"
    # Freeze header row for readability
    sheet.freeze_panes = "A2"

    headers = ["ID", "Title", "Vote Average", "Genres"]
    sheet.append(headers)

    bold_font = Font(bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for movie in rows:
        genres_text = ", ".join(movie.genres)
        sheet.append([movie.id, movie.title, movie.vote_average, genres_text])

        if movie.is_action:
            current_row = sheet.max_row
            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=current_row, column=col)
                cell.font = bold_font
                cell.fill = red_fill

    temp_path = f"{output_path}.tmp"
    try:
        workbook.save(temp_path)
        os.replace(temp_path, output_path)
    except (PermissionError, IOError, IllegalCharacterError) as e:
        logging.error(f"Failed to save Excel file due to an OS or data error: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)


def write_failed_ids(
    failed_ids: List[int],
    skipped_rows: List[Tuple[str, int, str]],
    output_path: str,
) -> None:
    """Write failed/skipped movie IDs to a CSV dead letter queue."""
    if not failed_ids and not skipped_rows:
        return

    try:
        with open(output_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                ["ID", "Reason", "RowNumber"]
            )  # RowNumber only for input skips
            for mid in sorted(failed_ids):
                writer.writerow(
                    [mid, "fetch-failed", ""]
                )  # no input row for API failures
            for raw_value, row_number, reason in skipped_rows:
                writer.writerow([raw_value, reason, row_number])
        logging.info(f"Wrote failed/skipped IDs to {output_path}")
    except OSError as e:
        logging.error(f"Failed to write dead letter queue: {e}")
